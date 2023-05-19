import os

from openpyxl.reader.excel import load_workbook
from rest_framework import serializers
from .models import Places


class PlaceSerializer(serializers.ModelSerializer):
    _TARGET_TYPE = '.xlsx'
    _MAX_COLUMNS = 3
    _START_HEADER_IDX = 1
    _END_HEADER_IDX = 4
    _HEADERS_NAMING = ['title', 'place', 'rating']

    id = serializers.PrimaryKeyRelatedField(
        read_only=True
    )

    title = serializers.CharField(
        read_only=True
    )

    rating = serializers.IntegerField(
        read_only=True
    )

    latitude = serializers.FloatField(
        read_only=True
    )

    longitude = serializers.FloatField(
        read_only=True
    )

    xlsx_file = serializers.FileField(
        write_only=True
    )

    def validate(self, attrs):
        def is_header_valid(header: str, column_num: int) -> bool:
            if header.lower() == self._HEADERS_NAMING[column_num - 1]:
                return True
            return False

        xlsx_file = attrs['xlsx_file']
        name, extension = os.path.splitext(xlsx_file.name)
        if extension != self._TARGET_TYPE:
            raise serializers.ValidationError(
                {'xlsx_file': f'{extension} type is incorrect type, {self._TARGET_TYPE} require'})

        workbook = load_workbook(filename=xlsx_file.file)
        worksheet = workbook.active

        if worksheet.max_column != self._MAX_COLUMNS:
            raise serializers.ValidationError({'xlsx_file': 'Must contain 3 columns'})

        for col in range(self._START_HEADER_IDX, self._END_HEADER_IDX):
            col_header = worksheet.cell(1, col)
            if not is_header_valid(col_header.value, col):
                raise serializers.ValidationError(
                    {
                        'xlsx_file': f'Field {col_header.value} does not supported. Available: {self._HEADERS_NAMING}'
                    })

        return attrs

    class Meta:
        model = Places
        fields = (
            'id',
            'title',
            'rating',
            'latitude',
            'longitude',
            'xlsx_file'
        )
