import os
from rest_framework import serializers
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from django.utils.deconstruct import deconstructible


@deconstructible
class XlsxValidator:
    _TARGET_TYPE = '.xlsx'
    _MAX_COLUMNS = 3
    _START_HEADER_IDX = 1
    _END_HEADER_IDX = 4
    _HEADERS_NAMING = ['title', 'place', 'rating']
    _POINT_DIVIDER = ','
    _POINT_COORDINATES_COUNT = 2
    _LAT_MIN_BORDER = -90.0
    _LAT_MAX_BORDER = 90.0
    _LONG_MIN_BORDER = -180.0
    _LONG_MAX_BORDER = 180.0

    def _is_header_valid(self, header: str, column_num: int) -> bool:
        if header.lower() == self._HEADERS_NAMING[column_num - 1]:
            return True
        return False

    def _validate_extension(self, file_name: str):
        name, extension = os.path.splitext(file_name)
        if extension != self._TARGET_TYPE:
            raise serializers.ValidationError(
                {
                    'xlsx_file': f'{extension} type is incorrect type, {self._TARGET_TYPE} require'
                })

    def _validate_fields_name(self, worksheet: Worksheet):
        for col in range(self._START_HEADER_IDX, self._END_HEADER_IDX):
            col_header = worksheet.cell(1, col)
            if not self._is_header_valid(col_header.value, col):
                raise serializers.ValidationError(
                    {
                        'xlsx_file': f'Field {col_header.value} does not supported. Available: {self._HEADERS_NAMING}'
                    })

    def _validate_cells(self, worksheet: Worksheet):
        def _validate_title(title: str):
            if not title:
                raise serializers.ValidationError(
                    {
                        'title': 'Title can not be empty'
                    })

        def _validate_point(point_string: str):
            if not point_string:
                raise serializers.ValidationError(
                    {
                        'point': 'Point can not be empty'
                    })

            if self._POINT_DIVIDER not in point_string:
                raise serializers.ValidationError(
                    {
                        'point': f'Point must contain "{self._POINT_DIVIDER}" as divider between lat and long'
                    })

            split_point_string = point_string.split(self._POINT_DIVIDER)
            if len(split_point_string) != self._POINT_COORDINATES_COUNT:
                raise serializers.ValidationError(
                    {
                        'point': f'Point must contain only {self._POINT_COORDINATES_COUNT} coordinates'
                    })
            try:
                lat_long = float(split_point_string[0]), float(split_point_string[1])
            except ValueError:
                raise serializers.ValidationError(
                    {
                        'point': f'Unable cast to float coordinates {split_point_string[0]}, {split_point_string[1]}'
                    })

            lat, long = lat_long
            if not self._LAT_MIN_BORDER <= lat <= self._LAT_MAX_BORDER:
                raise serializers.ValidationError(
                    {
                        'point': f'Point latitude {lat} not in interval [{self._LAT_MIN_BORDER}:{self._LAT_MAX_BORDER}]'
                    })

            if not self._LONG_MIN_BORDER <= long <= self._LONG_MAX_BORDER:
                raise serializers.ValidationError(
                    {
                        'point': f'Point longitude {long} not in interval [{self._LONG_MIN_BORDER}:{self._LONG_MAX_BORDER}]'
                    })

        def _validate_rating(rating_string: str):
            try:
                _ = int(rating_string)
            except ValueError:
                raise serializers.ValidationError(
                    {
                        'rating': f'Rating value {rating_string} can not be converted to int'
                    })

        for row in range(self._START_HEADER_IDX, worksheet.max_row):
            col_idx = 0
            for col in worksheet.iter_cols(self._START_HEADER_IDX, worksheet.max_column):
                value = col[row].value
                if col_idx == 0:
                    _validate_title(value)
                if col_idx == 1:
                    _validate_point(value)
                if col_idx == 3:
                    _validate_rating(value)
                col_idx += 1

    def __call__(self, value):
        xlsx_file = value['xlsx_file']

        self._validate_extension(xlsx_file.name)
        workbook = load_workbook(filename=xlsx_file.file)
        worksheet = workbook.active

        self._validate_fields_name(worksheet)
        self._validate_cells(worksheet)
