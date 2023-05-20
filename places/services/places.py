from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from places.models import Places

_PLACE_COL_IDX = 1
_START_ROW_IDX = 1
_START_COL_IDX = 1


def _parse_row(worksheet: Worksheet, row: int) -> Places:
    def type_cast(string_data: str, column_index: int):
        """
        Converts XLSX string type by current column to
        the data type needed for Places
        """
        result = string_data
        if column_index == _PLACE_COL_IDX:
            result = string_data.split(',')
            result = (float(result[0]), float(result[1]))
        return result

    col_idx = 0
    params = []
    for col in worksheet.iter_cols(_START_COL_IDX, worksheet.max_column):
        value = col[row].value
        value = type_cast(value, col_idx)
        params.append(value)

    return Places.objects.create(
        title=params[0],
        latitude=params[1][0],
        longitude=params[1][1],
        rating=params[2]
    )


def create_places(worksheet: Worksheet) -> List[Places]:
    places = []
    for i in range(_START_ROW_IDX, worksheet.max_row):
        place = _parse_row(worksheet, i)
        places.append(place)
    return places
